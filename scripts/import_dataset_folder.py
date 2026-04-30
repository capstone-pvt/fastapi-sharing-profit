"""Import the capstone dataset folder into MongoDB.

Reads from `<repo>/dataset/`:
  <species>/                                  — extra classifier images
  Individual/<species>/<label>.jpg            — per-fish images, mapped to Excel
  Tub/<species>/<label>(_coin)?.jpg           — per-tub images, mapped to Excel
  species weight.xlsx                         — ground-truth weights
  Fish Weight Estimation Dataset.xlsx         — ground-truth (older copy)
  Fish Weight Estimation Dataset_NEW.xlsx     — ground-truth (newer copy)

  Excel sheets used:
    Individual: fish_length(cm), fish_grid(cm), actual_weight(kilos)
    Tub:        tub_length(cm),  tub_width(cm), actual_weight(kg)

Writes (after `clear_training_data.py` has emptied them):
  fish_species             — one doc per species (folder name), classIndex 0..N-1
  fish_training_samples    — one doc per discovered image:
       - source="dataset_import" for classifier-only samples (no weight)
       - source="excel_individual" / "excel_tub" for Excel-matched rows with
         weightKg + lengthCm + widthCm populated. The Excel `image` column is
         matched to the actual file in Individual/<species>/ or Tub/<species>/.

Image bytes are not copied. `imagePath` references the dataset folder so the
existing exporter (`infrastructure/fish_training_samples/exporter.py`) reads
files in place.

Usage:
    cd profit_sharing_api_fastapi
    PYTHONPATH=. python scripts/import_dataset_folder.py
"""

from __future__ import annotations

import asyncio
import math
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
import openpyxl

_project_root = Path(__file__).resolve().parents[1]
load_dotenv(_project_root / ".env")

from app.db import connect_db, disconnect_db, get_db  # noqa: E402


DATASET_ROOT = _project_root.parent / "dataset"
EXCEL_FILES = [
    DATASET_ROOT / "species weight.xlsx",
    DATASET_ROOT / "Fish Weight Estimation Dataset.xlsx",
    DATASET_ROOT / "Fish Weight Estimation Dataset_NEW.xlsx",
]
INDIVIDUAL_ROOT = DATASET_ROOT / "Individual"
TUB_ROOT = DATASET_ROOT / "Tub"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".jfif"}

SPECIES_TAXONOMY: dict[str, dict[str, str]] = {
    "Auxis rochei": {
        "family": "Scombridae", "genus": "Auxis",
        "scientificName": "Auxis rochei",
        "englishName": "Bullet tuna", "localName": "Mangko/Pirit",
    },
    "Elagatis bipinnulata": {
        "family": "Carangidae", "genus": "Elagatis",
        "scientificName": "Elagatis bipinnulata",
        "englishName": "Rainbow runner", "localName": "Salindatu/Salmon",
    },
    "Euthynnus affinis": {
        "family": "Scombridae", "genus": "Euthynnus",
        "scientificName": "Euthynnus affinis",
        "englishName": "Eastern little tuna", "localName": "Patikan/Tulingan",
    },
    "Katsuwonus pelamis": {
        "family": "Scombridae", "genus": "Katsuwonus",
        "scientificName": "Katsuwonus pelamis",
        "englishName": "Skipjack tuna", "localName": "Sambagon/Tulingan/Bulis",
    },
    "Thunnus albacares": {
        "family": "Scombridae", "genus": "Thunnus",
        "scientificName": "Thunnus albacares",
        "englishName": "Yellowfin tuna", "localName": "Barilis/Bariles/Karaw",
    },
}


def _safe_float(value):
    if value is None or value == "":
        return None
    try:
        f = float(value)
        return f if math.isfinite(f) else None
    except (TypeError, ValueError):
        return None


def _list_species_folders() -> list[Path]:
    skip = {"Individual", "Tub", "Individual-20260429T051121Z-3-001"}
    return sorted(
        p for p in DATASET_ROOT.iterdir()
        if p.is_dir() and p.name not in skip
    )


def _list_images(folder: Path) -> list[Path]:
    if not folder.exists():
        return []
    return sorted(p for p in folder.iterdir() if p.suffix.lower() in IMAGE_EXTS)


def _gather_classifier_images(species_name: str) -> list[Path]:
    """Combine images from top-level, Individual/, and Tub/ folders."""
    images: list[Path] = []
    for root in (DATASET_ROOT / species_name,
                 INDIVIDUAL_ROOT / species_name,
                 TUB_ROOT / species_name):
        images.extend(_list_images(root))
    seen: set[str] = set()
    unique: list[Path] = []
    for p in images:
        key = str(p.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(p)
    return unique


def _resolve_image(folder: Path, label: str) -> Path | None:
    """Find a file in `folder` whose stem matches `label` (case-insensitive)."""
    if not folder.exists() or not label:
        return None
    target = label.strip().lower()
    for ext in (".jpg", ".jpeg", ".png", ".bmp", ".webp", ".jfif"):
        cand = folder / f"{label}{ext}"
        if cand.exists():
            return cand
    for p in folder.iterdir():
        if p.suffix.lower() in IMAGE_EXTS and p.stem.lower() == target:
            return p
    return None


async def _seed_species(folders: list[Path]) -> dict[str, dict]:
    db = get_db()
    now = datetime.now(timezone.utc)
    species_docs: dict[str, dict] = {}
    for class_index, folder in enumerate(folders):
        name = folder.name
        taxonomy = SPECIES_TAXONOMY.get(name, {})
        doc = {
            "name": name, "classIndex": class_index, "isActive": True,
            "scientificName": taxonomy.get("scientificName", name),
            "genus": taxonomy.get("genus"),
            "family": taxonomy.get("family"),
            "englishName": taxonomy.get("englishName"),
            "localName": taxonomy.get("localName"),
            "createdAt": now, "updatedAt": now,
        }
        result = await db["fish_species"].insert_one(doc)
        doc["_id"] = result.inserted_id
        species_docs[name] = doc
        print(f"  Species[{class_index}] {name}")
    return species_docs


def _build_classifier_doc(image_path: Path, species: dict, now) -> dict:
    return {
        "userId": None,
        "imageUrl": None,
        "imagePath": str(image_path.resolve()),
        "species": species["name"],
        "scientificName": species.get("scientificName"),
        "englishName": species.get("englishName"),
        "localName": species.get("localName"),
        "weightKg": None, "pricePerKg": None,
        "lengthCm": None, "widthCm": None, "heightCm": None,
        "scaleReferenceCm": None,
        "notes": "Imported from dataset folder",
        "bbox": None, "capturedAt": None,
        "createdAt": now, "updatedAt": now,
        "trainedBy": None,
        "source": "dataset_import",
        "originalSpecies": None, "originalConfidence": None, "analysisId": None,
    }


def _build_weight_doc(
    image_path: Path, species: dict, weightKg, lengthCm, widthCm, heightCm,
    label: str, sheet: str, now,
) -> dict:
    return {
        "userId": None,
        "imageUrl": None,
        "imagePath": str(image_path.resolve()),
        "species": species["name"],
        "scientificName": species.get("scientificName"),
        "englishName": species.get("englishName"),
        "localName": species.get("localName"),
        "weightKg": weightKg, "pricePerKg": None,
        "lengthCm": lengthCm, "widthCm": widthCm, "heightCm": heightCm,
        "scaleReferenceCm": None,
        "notes": f"Excel ground-truth ({sheet}): {label}",
        "bbox": None, "capturedAt": None,
        "createdAt": now, "updatedAt": now,
        "trainedBy": None,
        "source": f"excel_{sheet.lower()}",
        "originalSpecies": None, "originalConfidence": None, "analysisId": None,
    }


def _read_excel_rows() -> tuple[list[dict], list[dict]]:
    """Read + deduplicate Individual/Tub rows from all Excel files."""
    individual: dict[tuple, dict] = {}
    tub: dict[tuple, dict] = {}
    for path in EXCEL_FILES:
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, data_only=True)

        if "Individual" in wb.sheetnames:
            for r in list(wb["Individual"].iter_rows(values_only=True))[1:]:
                if not r or not r[0] or not r[6]:
                    continue
                key = (r[0], r[6])
                individual[key] = {
                    "species": r[0], "label": r[6],
                    "lengthCm": _safe_float(r[7]),
                    "widthCm": _safe_float(r[8]),
                    "weightKg": _safe_float(r[9]),
                    "sheet": "Individual",
                }

        if "Tub" in wb.sheetnames:
            for r in list(wb["Tub"].iter_rows(values_only=True))[1:]:
                if not r or not r[0] or not r[6]:
                    continue
                key = (r[0], r[6])
                tub[key] = {
                    "species": r[0],
                    "label_no_coin": r[6],
                    "label_with_coin": r[7],
                    "lengthCm": _safe_float(r[8]),
                    "widthCm": _safe_float(r[9]),
                    "heightCm": _safe_float(r[10]),
                    "weightKg": _safe_float(r[11]),
                    "sheet": "Tub",
                }

    return list(individual.values()), list(tub.values())


async def _seed_classifier_samples(
    folders: list[Path], species_docs: dict[str, dict]
) -> tuple[dict[str, list[Path]], int]:
    db = get_db()
    now = datetime.now(timezone.utc)
    by_species: dict[str, list[Path]] = {}
    docs: list[dict] = []
    for folder in folders:
        species_name = folder.name
        species = species_docs[species_name]
        images = _gather_classifier_images(species_name)
        by_species[species_name] = images
        for img in images:
            docs.append(_build_classifier_doc(img, species, now))
        print(f"  Images[{species_name}]: {len(images)}")
    if docs:
        await db["fish_training_samples"].insert_many(docs)
    return by_species, len(docs)


async def _seed_excel_weight_samples(
    individual_rows: list[dict],
    tub_rows: list[dict],
    species_docs: dict[str, dict],
) -> dict[str, int]:
    db = get_db()
    now = datetime.now(timezone.utc)
    docs: list[dict] = []

    indiv_matched = 0
    indiv_unmatched: list[str] = []
    for row in individual_rows:
        species_name = row["species"]
        species = species_docs.get(species_name)
        if species is None or row["weightKg"] is None:
            continue
        img = _resolve_image(INDIVIDUAL_ROOT / species_name, row["label"])
        if img is None:
            indiv_unmatched.append(f"{species_name}/{row['label']}")
            continue
        docs.append(
            _build_weight_doc(
                img, species, row["weightKg"],
                row["lengthCm"], row["widthCm"], None,
                row["label"], "Individual", now,
            )
        )
        indiv_matched += 1

    tub_matched = 0
    tub_unmatched: list[str] = []
    for row in tub_rows:
        species_name = row["species"]
        species = species_docs.get(species_name)
        if species is None or row["weightKg"] is None:
            continue
        # Prefer the without-coin photo for the weight regression sample,
        # fall back to the with-coin variant if only that exists.
        img = (
            _resolve_image(TUB_ROOT / species_name, row["label_no_coin"]) or
            _resolve_image(TUB_ROOT / species_name, row["label_with_coin"])
        )
        if img is None:
            tub_unmatched.append(f"{species_name}/{row['label_no_coin']}")
            continue
        docs.append(
            _build_weight_doc(
                img, species, row["weightKg"],
                row["lengthCm"], row["widthCm"], row.get("heightCm"),
                row["label_no_coin"], "Tub", now,
            )
        )
        tub_matched += 1

    if docs:
        await db["fish_training_samples"].insert_many(docs)

    print(f"  Individual rows matched: {indiv_matched} / {len(individual_rows)}")
    if indiv_unmatched:
        print(f"    Unmatched: {indiv_unmatched}")
    print(f"  Tub        rows matched: {tub_matched} / {len(tub_rows)}")
    if tub_unmatched:
        print(f"    Unmatched: {tub_unmatched}")

    return {"individual": indiv_matched, "tub": tub_matched}


async def main() -> None:
    if not DATASET_ROOT.exists():
        raise SystemExit(f"Dataset root not found: {DATASET_ROOT}")

    folders = _list_species_folders()
    if not folders:
        raise SystemExit(f"No species folders found in {DATASET_ROOT}")

    print("=" * 64)
    print("  SEEDING SPECIES")
    print("=" * 64)
    await connect_db()
    try:
        species_docs = await _seed_species(folders)

        print()
        print("=" * 64)
        print("  IMPORTING CLASSIFIER IMAGES")
        print("=" * 64)
        by_species, total_class = await _seed_classifier_samples(folders, species_docs)

        print()
        print("=" * 64)
        print("  IMPORTING EXCEL WEIGHT ROWS")
        print("=" * 64)
        individual_rows, tub_rows = _read_excel_rows()
        excel_counts = await _seed_excel_weight_samples(
            individual_rows, tub_rows, species_docs
        )

        print()
        print("=" * 64)
        print("  SUMMARY")
        print("=" * 64)
        print(f"  Species:                  {len(species_docs)}")
        print(f"  Classifier-only samples:  {total_class}")
        print(
            f"  Excel weighted samples:   "
            f"{excel_counts['individual'] + excel_counts['tub']} "
            f"(Individual={excel_counts['individual']}, Tub={excel_counts['tub']})"
        )
        print(
            f"  Total training samples:   "
            f"{total_class + excel_counts['individual'] + excel_counts['tub']}"
        )
    finally:
        await disconnect_db()

    print()
    print("Next: run training with")
    print("  PATH=$(pwd)/venv/Scripts:$PATH PYTHONPATH=. \\")
    print("    ./venv/Scripts/python.exe scripts/auto_train.py --skip-detect "
          "--epochs-classify 25 --imgsz 224")


if __name__ == "__main__":
    asyncio.run(main())
