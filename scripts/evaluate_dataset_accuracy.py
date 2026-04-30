"""Evaluate species classifier and weight regressor on the dataset folder.

Splits weight metrics by Excel sheet (Individual vs Tub) so each context is
reported separately. Species accuracy is reported overall and per-species
using the val split that auto_train.py created.

Usage:
    cd profit_sharing_api_fastapi
    PYTHONPATH=. python scripts/evaluate_dataset_accuracy.py
"""

from __future__ import annotations

import math
from pathlib import Path

import joblib
import numpy as np
import openpyxl
from PIL import Image, ImageEnhance, ImageOps

_project_root = Path(__file__).resolve().parents[1]
DATASET_ROOT = _project_root.parent / "dataset"
EXCEL_FILE = DATASET_ROOT / "Fish Weight Estimation Dataset.xlsx"
CLASSIFIER_PATH = _project_root / "app" / "models" / "classifier" / "best.pt"
WEIGHT_MODEL_PATH = _project_root / "app" / "models" / "weight" / "weight_model.joblib"
VAL_DIR = _project_root / "datasets" / "fish_species" / "val"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".jfif"}


def _safe_float(value):
    if value is None or value == "":
        return None
    try:
        f = float(value)
        return f if math.isfinite(f) else None
    except (TypeError, ValueError):
        return None


def _read_excel():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    individual = []
    for r in list(wb["Individual"].iter_rows(values_only=True))[1:]:
        if not r or not r[0]:
            continue
        individual.append(
            {
                "species": r[0],
                "label": r[6],
                "lengthCm": _safe_float(r[7]),
                "widthCm": _safe_float(r[8]),
                "weightKg": _safe_float(r[9]),
            }
        )

    tub = []
    for r in list(wb["Tub"].iter_rows(values_only=True))[1:]:
        if not r or not r[0]:
            continue
        tub.append(
            {
                "species": r[0],
                "label": r[6],
                "lengthCm": _safe_float(r[8]),
                "widthCm": _safe_float(r[9]),
                "heightCm": _safe_float(r[10]),
                "weightKg": _safe_float(r[11]),
            }
        )

    return individual, tub


def _build_species_index(folders: list[Path]) -> dict[str, int]:
    return {f.name: i for i, f in enumerate(sorted(folders))}


def _add_engineered(X: np.ndarray) -> np.ndarray:
    """7 base + 3 engineered. Must match train_regression_models.py."""
    bbox_w = X[:, 1]
    bbox_h = X[:, 2]
    length_cm = X[:, 4]
    width_cm = X[:, 5]
    height_cm = X[:, 6]
    bbox_area = (bbox_w * bbox_h).reshape(-1, 1)
    bbox_aspect = np.where(bbox_h > 0, bbox_w / np.maximum(bbox_h, 1.0), 1.0).reshape(-1, 1)
    volume_cm = (length_cm * width_cm * height_cm).reshape(-1, 1)
    return np.hstack([X, bbox_area, bbox_aspect, volume_cm])


def _eval_weight(rows: list[dict], species_index: dict[str, int], model) -> dict:
    if not rows:
        return {"n": 0}

    valid = [r for r in rows if r["weightKg"] and r["lengthCm"] and r["widthCm"]]
    if not valid:
        return {"n": 0}

    X_rows = []
    y_true = []
    for r in valid:
        idx = species_index.get(r["species"], -1)
        X_rows.append([
            idx, 0.0, 0.0, 0.0,
            r["lengthCm"], r["widthCm"],
            float(r.get("heightCm") or 0),
        ])
        y_true.append(r["weightKg"])

    X = _add_engineered(np.array(X_rows, dtype=float))
    y_true = np.array(y_true, dtype=float)
    y_pred = model.predict(X)

    abs_err = np.abs(y_pred - y_true)
    pct_err = abs_err / np.maximum(y_true, 1e-6)
    mae = float(abs_err.mean())
    rmse = float(np.sqrt(((y_pred - y_true) ** 2).mean()))
    mape = float(pct_err.mean())
    ss_res = float(((y_true - y_pred) ** 2).sum())
    ss_tot = float(((y_true - y_true.mean()) ** 2).sum())
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    within_10 = float((pct_err <= 0.10).mean())
    within_20 = float((pct_err <= 0.20).mean())

    per_row = []
    for r, true_w, pred_w in zip(valid, y_true, y_pred):
        per_row.append(
            {
                "species": r["species"],
                "label": r["label"],
                "lengthCm": r["lengthCm"],
                "widthCm": r["widthCm"],
                "trueKg": float(true_w),
                "predKg": float(pred_w),
                "absErr": float(abs(pred_w - true_w)),
                "pctErr": float(abs(pred_w - true_w) / max(true_w, 1e-6)),
            }
        )

    return {
        "n": len(valid),
        "mae_kg": mae,
        "rmse_kg": rmse,
        "mape": mape,
        "r2": r2,
        "within_10_pct": within_10,
        "within_20_pct": within_20,
        "rows": per_row,
    }


def _eval_classifier(model_path: Path) -> dict:
    from ultralytics import YOLO

    if not VAL_DIR.exists():
        return {"error": f"Val dir not found: {VAL_DIR}"}

    model = YOLO(str(model_path))

    per_class = {}
    confusion = {}
    total = 0
    correct = 0

    for class_dir in sorted(VAL_DIR.iterdir()):
        if not class_dir.is_dir():
            continue
        true_class = class_dir.name
        images = [p for p in class_dir.iterdir() if p.suffix.lower() in IMAGE_EXTS]
        c_total = 0
        c_correct = 0
        for img_path in images:
            try:
                pil_img = Image.open(img_path).convert("RGB")
                pil_img = ImageOps.exif_transpose(pil_img) or pil_img
                pil_img = ImageEnhance.Sharpness(pil_img).enhance(1.3)
                pil_img = ImageEnhance.Contrast(pil_img).enhance(1.1)

                results = model.predict(pil_img, verbose=False)
                if not results or not hasattr(results[0], "probs"):
                    continue

                probs = results[0].probs.data.cpu().numpy()
                names = results[0].names
                pred_idx = int(np.argmax(probs))
                pred_class = names.get(pred_idx, "Unknown")

                c_total += 1
                total += 1
                if pred_class.lower() == true_class.lower():
                    c_correct += 1
                    correct += 1
                confusion.setdefault(true_class, {}).setdefault(pred_class, 0)
                confusion[true_class][pred_class] += 1
            except Exception as e:
                print(f"  Warning: {img_path.name}: {e}")

        if c_total > 0:
            per_class[true_class] = {
                "n": c_total,
                "correct": c_correct,
                "accuracy": c_correct / c_total,
            }

    return {
        "total": total,
        "correct": correct,
        "accuracy": correct / total if total else 0.0,
        "per_class": per_class,
        "confusion": confusion,
    }


def _print_weight_summary(name: str, result: dict) -> None:
    print()
    print("-" * 64)
    print(f"  WEIGHT REGRESSION: {name}")
    print("-" * 64)
    if result.get("n", 0) == 0:
        print("  No usable rows.")
        return

    print(f"  Samples evaluated:        {result['n']}")
    print(f"  MAE  (mean abs error):    {result['mae_kg']:.3f} kg")
    print(f"  RMSE (root mean sq err):  {result['rmse_kg']:.3f} kg")
    print(f"  MAPE (mean abs % err):    {result['mape']:.1%}")
    print(f"  R^2:                      {result['r2']:.3f}")
    print(f"  Within +/-10% of truth:   {result['within_10_pct']:.1%}")
    print(f"  Within +/-20% of truth:   {result['within_20_pct']:.1%}")

    by_species = {}
    for row in result["rows"]:
        by_species.setdefault(row["species"], []).append(row)
    print()
    print(f"  Per-species:")
    for species, rows in sorted(by_species.items()):
        errs = [r["absErr"] for r in rows]
        pcts = [r["pctErr"] for r in rows]
        print(
            f"    {species:25s} n={len(rows):3d}  "
            f"MAE={sum(errs)/len(errs):6.3f} kg  "
            f"MAPE={sum(pcts)/len(pcts):6.1%}"
        )

    print()
    print(f"  Worst predictions:")
    worst = sorted(result["rows"], key=lambda r: -r["pctErr"])[:5]
    for r in worst:
        print(
            f"    {r['species']:22s} {r['label']:18s} "
            f"true={r['trueKg']:6.2f} kg  pred={r['predKg']:6.2f} kg  "
            f"err={r['pctErr']:6.1%}"
        )


def _print_classifier_summary(result: dict) -> None:
    print()
    print("=" * 64)
    print("  SPECIES CLASSIFICATION (val split)")
    print("=" * 64)
    if "error" in result:
        print(f"  {result['error']}")
        return
    print(f"  Total val images:    {result['total']}")
    print(
        f"  Top-1 accuracy:      {result['accuracy']:.1%} "
        f"({result['correct']}/{result['total']})"
    )
    print()
    print(f"  Per-species:")
    for species, stats in sorted(result["per_class"].items()):
        bar = "PASS" if stats["accuracy"] >= 0.9 else (
            "WARN" if stats["accuracy"] >= 0.7 else "FAIL"
        )
        print(
            f"    [{bar}] {species:25s} {stats['accuracy']:6.1%}  "
            f"({stats['correct']}/{stats['n']})"
        )

    misses = []
    for true_cls, preds in result["confusion"].items():
        for pred_cls, n in preds.items():
            if true_cls.lower() != pred_cls.lower():
                misses.append((true_cls, pred_cls, n))
    misses.sort(key=lambda t: -t[2])
    if misses:
        print()
        print(f"  Top misclassifications:")
        for t, p, n in misses[:10]:
            print(f"    {t} -> {p}: {n}")


def main() -> None:
    if not WEIGHT_MODEL_PATH.exists():
        raise SystemExit(f"Weight model not found: {WEIGHT_MODEL_PATH}")
    if not CLASSIFIER_PATH.exists():
        raise SystemExit(f"Classifier model not found: {CLASSIFIER_PATH}")
    if not EXCEL_FILE.exists():
        raise SystemExit(f"Excel file not found: {EXCEL_FILE}")

    skip = {"Individual", "Tub", "Individual-20260429T051121Z-3-001"}
    folders = sorted(
        p for p in DATASET_ROOT.iterdir() if p.is_dir() and p.name not in skip
    )
    species_index = _build_species_index(folders)

    print("=" * 64)
    print("  EVALUATION")
    print("=" * 64)
    print(f"  Classifier:    {CLASSIFIER_PATH}")
    print(f"  Weight model:  {WEIGHT_MODEL_PATH}")
    print(f"  Excel:         {EXCEL_FILE.name}")
    print(f"  Species index: {species_index}")

    weight_model = joblib.load(WEIGHT_MODEL_PATH)

    individual_rows, tub_rows = _read_excel()
    indiv_result = _eval_weight(individual_rows, species_index, weight_model)
    tub_result = _eval_weight(tub_rows, species_index, weight_model)

    _print_weight_summary("Individual sheet", indiv_result)
    _print_weight_summary("Tub sheet", tub_result)

    cls_result = _eval_classifier(CLASSIFIER_PATH)
    _print_classifier_summary(cls_result)

    print()
    print("=" * 64)
    print("  DONE")
    print("=" * 64)


if __name__ == "__main__":
    main()
