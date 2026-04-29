"""Honest weight-regressor accuracy via stratified k-fold cross-validation.

Reads the same Excel rows used to train the weight model, but trains and
predicts in folds so every row's predicted weight comes from a model that
*didn't* see it during training. Reports MAE / RMSE / MAPE / R² split by
sheet (Individual vs Tub) and per species.

Why
  scripts/evaluate_dataset_accuracy.py reports training-set numbers, which
  flatter the model. With only 54 rows there's no luxury of a permanent
  hold-out, so this script uses K-fold CV (default K=5) and aggregates the
  out-of-fold predictions to produce one honest prediction per row.

Usage
    cd profit_sharing_api_fastapi
    PYTHONPATH=. python scripts/evaluate_weight_holdout.py [--folds 5] [--seed 42]
"""

from __future__ import annotations

import argparse
import math
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.model_selection import StratifiedKFold
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

_project_root = Path(__file__).resolve().parents[1]
DATASET_ROOT = _project_root.parent / "dataset"
EXCEL_FILES = [
    DATASET_ROOT / "species weight.xlsx",
    DATASET_ROOT / "Fish Weight Estimation Dataset.xlsx",
    DATASET_ROOT / "Fish Weight Estimation Dataset_NEW.xlsx",
]


def _safe_float(value):
    if value is None or value == "":
        return None
    try:
        f = float(value)
        return f if math.isfinite(f) else None
    except (TypeError, ValueError):
        return None


def _read_rows() -> list[dict]:
    individual: dict[tuple, dict] = {}
    tub: dict[tuple, dict] = {}
    for path in EXCEL_FILES:
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Individual" in wb.sheetnames:
            for r in list(wb["Individual"].iter_rows(values_only=True))[1:]:
                if not r or not r[0] or not r[6]:
                    continue
                individual[(r[0], r[6])] = {
                    "sheet": "Individual",
                    "species": r[0], "label": r[6],
                    "lengthCm": _safe_float(r[7]),
                    "widthCm": _safe_float(r[8]),
                    "weightKg": _safe_float(r[9]),
                }
        if "Tub" in wb.sheetnames:
            for r in list(wb["Tub"].iter_rows(values_only=True))[1:]:
                if not r or not r[0] or not r[6]:
                    continue
                tub[(r[0], r[6])] = {
                    "sheet": "Tub",
                    "species": r[0], "label": r[6],
                    "lengthCm": _safe_float(r[8]),
                    "widthCm": _safe_float(r[9]),
                    "weightKg": _safe_float(r[11]),
                }
    rows = list(individual.values()) + list(tub.values())
    rows = [
        r for r in rows
        if r["weightKg"] and r["weightKg"] > 0
        and r["lengthCm"] and r["widthCm"]
    ]
    return rows


def _build_features(rows: list[dict], species_index: dict[str, int]) -> tuple[np.ndarray, np.ndarray]:
    """Same feature contract as scripts/train_regression_models.py."""
    X_rows = []
    y = []
    for r in rows:
        idx = species_index[r["species"]]
        X_rows.append([
            float(idx),
            0.0,                # bboxWidth (no detector data for Excel rows)
            0.0,                # bboxHeight
            0.0,                # scaleReferenceCm
            float(r["lengthCm"]),
            float(r["widthCm"]),
        ])
        y.append(float(r["weightKg"]))
    X = np.array(X_rows, dtype=float)
    bbox_w = X[:, 1]
    bbox_h = X[:, 2]
    length_cm = X[:, 4]
    width_cm = X[:, 5]
    area = (bbox_w * bbox_h).reshape(-1, 1)
    aspect = np.where(bbox_h > 0, bbox_w / np.maximum(bbox_h, 1.0), 1.0).reshape(-1, 1)
    area_cm = (length_cm * width_cm).reshape(-1, 1)
    X_full = np.hstack([X, area, aspect, area_cm])
    return X_full, np.array(y, dtype=float)


def _make_model() -> Pipeline:
    return Pipeline([
        ("scaler", StandardScaler()),
        ("regressor", GradientBoostingRegressor(
            n_estimators=200, max_depth=4, learning_rate=0.1,
            subsample=0.8, random_state=42,
        )),
    ])


def _metrics(y_true: np.ndarray, y_pred: np.ndarray) -> dict:
    abs_err = np.abs(y_pred - y_true)
    pct_err = abs_err / np.maximum(y_true, 1e-6)
    mae = float(abs_err.mean())
    rmse = float(np.sqrt(((y_pred - y_true) ** 2).mean()))
    mape = float(pct_err.mean())
    ss_res = float(((y_true - y_pred) ** 2).sum())
    ss_tot = float(((y_true - y_true.mean()) ** 2).sum())
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    return {
        "n": len(y_true), "mae": mae, "rmse": rmse, "mape": mape, "r2": r2,
        "within_10": float((pct_err <= 0.10).mean()),
        "within_20": float((pct_err <= 0.20).mean()),
    }


def _print_metrics(label: str, m: dict) -> None:
    print(f"  {label}")
    print(f"    n:                    {m['n']}")
    print(f"    MAE  (kg):            {m['mae']:.3f}")
    print(f"    RMSE (kg):            {m['rmse']:.3f}")
    print(f"    MAPE:                 {m['mape']:.1%}")
    print(f"    R^2:                  {m['r2']:.3f}")
    print(f"    Within +/-10% truth:  {m['within_10']:.1%}")
    print(f"    Within +/-20% truth:  {m['within_20']:.1%}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Honest weight-regressor accuracy via k-fold CV."
    )
    parser.add_argument("--folds", type=int, default=5)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    rows = _read_rows()
    if not rows:
        raise SystemExit("No usable Excel rows found.")

    species_names = sorted({r["species"] for r in rows})
    species_index = {name: i for i, name in enumerate(species_names)}

    X, y = _build_features(rows, species_index)

    # Stratify by (sheet, species) so each fold has both regimes
    strat_labels = [f"{r['sheet']}|{r['species']}" for r in rows]
    label_counts: dict[str, int] = defaultdict(int)
    for s in strat_labels:
        label_counts[s] += 1
    # Drop labels that can't satisfy K-fold (need >= folds members)
    skf_labels = [s if label_counts[s] >= args.folds else "rare" for s in strat_labels]

    print("=" * 64)
    print(f"  HONEST WEIGHT EVAL — {args.folds}-FOLD CV")
    print("=" * 64)
    print(f"  Total rows:           {len(rows)}")
    print(f"  Species:              {species_names}")
    print(
        f"  Per-stratum sample sizes: "
        f"{dict(sorted(label_counts.items()))}"
    )
    print()

    skf = StratifiedKFold(n_splits=args.folds, shuffle=True, random_state=args.seed)
    oof_pred = np.full_like(y, fill_value=np.nan)

    for fold, (train_idx, test_idx) in enumerate(skf.split(X, skf_labels), start=1):
        model = _make_model()
        model.fit(X[train_idx], y[train_idx])
        oof_pred[test_idx] = model.predict(X[test_idx])
        fold_m = _metrics(y[test_idx], oof_pred[test_idx])
        print(
            f"  Fold {fold}/{args.folds}: "
            f"n={fold_m['n']:2d}  MAE={fold_m['mae']:.3f} kg  "
            f"MAPE={fold_m['mape']:.1%}  R²={fold_m['r2']:6.3f}"
        )

    print()
    print("-" * 64)
    print("  OVERALL HOLDOUT METRICS (one prediction per row, model trained without it)")
    print("-" * 64)
    _print_metrics("All 54 rows:", _metrics(y, oof_pred))

    print()
    print("-" * 64)
    print("  BY SHEET")
    print("-" * 64)
    for sheet in ("Individual", "Tub"):
        mask = np.array([r["sheet"] == sheet for r in rows])
        if not mask.any():
            continue
        _print_metrics(f"{sheet}:", _metrics(y[mask], oof_pred[mask]))
        print()
        # per-species inside the sheet
        for species in species_names:
            sm = mask & np.array([r["species"] == species for r in rows])
            if not sm.any():
                continue
            m = _metrics(y[sm], oof_pred[sm])
            print(
                f"      {species:25s}  n={m['n']:2d}  "
                f"MAE={m['mae']:.3f} kg  MAPE={m['mape']:6.1%}  "
                f"within±20%={m['within_20']:.0%}"
            )
        print()

    print("-" * 64)
    print("  WORST 8 ROWS (largest absolute % error)")
    print("-" * 64)
    rows_with_pred = sorted(
        zip(rows, y, oof_pred),
        key=lambda t: -abs(t[2] - t[1]) / max(t[1], 1e-6),
    )[:8]
    for r, true_kg, pred_kg in rows_with_pred:
        pct = abs(pred_kg - true_kg) / max(true_kg, 1e-6)
        print(
            f"    [{r['sheet']:10s}] {r['species']:22s} "
            f"{r['label']:20s} L={r['lengthCm']:5.1f} W={r['widthCm']:5.1f}  "
            f"true={true_kg:6.2f} kg  pred={pred_kg:6.2f} kg  err={pct:6.1%}"
        )

    print()
    print("=" * 64)
    print("  DONE")
    print("=" * 64)
    print(
        "  These numbers come from out-of-fold predictions, so each row's "
        "prediction was made by a model that did NOT see that row during "
        "training. This is the honest generalization signal — quote these, "
        "not the training-set numbers in evaluate_dataset_accuracy.py."
    )


if __name__ == "__main__":
    main()
